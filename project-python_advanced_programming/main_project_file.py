"""This file reads the data from the excel file which contains all 5 sheets
    Then that data is printed on the console after receiving the input of PS number
    from the user
    The same data is then copies into a new excel file
"""
import sys

try:
    import openpyxl
except ModuleNotFoundError as error:
    print(error)
    print("Please Install openpyxl module to use this project")


def get_ps_numbers(input_data_workbook):
    """Function to return list of ps numbers from the excel file"""
    list_of_ps_numbers = []
    sheet_names = input_data_workbook.sheetnames
    sheet1 = input_data_workbook[sheet_names[0]]
    max_rows = len(sheet1["A"])
    for row in range(2, max_rows + 1):
        list_of_ps_numbers.append(sheet1["A" + str(row)].value)
    return list_of_ps_numbers


def get_marks_ps_number(ps_number, input_data_workbook):
    """Function to return marks of a particular PS number
        from the given excel file as a list"""
    sheet_marks = input_data_workbook['Semester Marks Sheet']
    max_cols = len(sheet_marks["1"])
    semester_marks = []
    for col in range(66, 66 + max_cols - 1):
        cell_pos = chr(col) + str(ps_number)
        semester_marks.append(sheet_marks[cell_pos].value)
    return semester_marks


def get_hobbies_ps_number(ps_number, input_data_workbook):
    """Function to return Hobbies of a particular PS number
        from the given excel file as a list"""
    sheet_marks = input_data_workbook['Random Hobbies Sheet']
    max_cols = len(sheet_marks["1"])
    hobbies = []
    for col in range(66, 66 + max_cols - 1):
        cell_pos = chr(col) + str(ps_number)
        hobbies.append(sheet_marks[cell_pos].value)
    return hobbies


def get_cities_ps_number(ps_number, input_data_workbook):
    """Function to return Cities of a particular PS number
        from the given excel file as a list"""
    sheet_marks = input_data_workbook['Cities Visited Sheet']
    max_cols = len(sheet_marks["1"])
    cities = []
    for col in range(66, 66 + max_cols - 1):
        cell_pos = chr(col) + str(ps_number)
        cities.append(sheet_marks[cell_pos].value)
    return cities


def get_programming_ps_number(ps_number, input_data_workbook):
    """Function to return Programming skills of a
        particular PS number from the given excel
        file as a list"""
    sheet_marks = input_data_workbook['Programming Expertise Sheet']
    max_cols = len(sheet_marks["1"])
    programming_skills = []
    for col in range(66, 66 + max_cols - 1):
        cell_pos = chr(col) + str(ps_number)
        programming_skills.append(sheet_marks[cell_pos].value)
    return programming_skills


def get_domain_areas_ps_number(ps_number, input_data_workbook):
    """Function to return Domain Areas skills of a
        particular PS number from the given excel
        file as a list"""
    sheet_marks = input_data_workbook['Domain Areas Sheet']
    max_cols = len(sheet_marks["1"])
    domain_areas = []
    for col in range(66, 66 + max_cols - 1):
        cell_pos = chr(col) + str(ps_number)
        domain_areas.append(sheet_marks[cell_pos].value)
    return domain_areas


def print_welcome_screen(ps_numbers):
    """Function to print welcome screen and all PS numbers
    and ask user for PS number of his choice"""
    print("List Of available PS Numbers: ")
    sr_no = 1
    for number in ps_numbers:
        print(str(sr_no) + '.', number)
        sr_no += 1
    print("16. To exit the program")
    while True:
        try:
            user_choice = int(input("\nSelect the number you want to choose: ")) + 1
            if not 1 < user_choice < 18:
                print("Please Enter valid PS Number choice")
                print("Enter number between 1 to 15")
            elif user_choice == 17:
                print("Program Terminated....")
                sys.exit()
            else:
                return user_choice
        except ValueError:
            print("ERROR!! Enter only Integer")
            print("Enter a Integer number between 1-15")


def print_data_of_ps_number(ps_number, size_of_list):
    """Printing the data of the requested PS number as a tabular format
    on the console using the lists extracted from the excel file"""
    print(f"Data Requested by the user for the PS"
          f" Number {ps_number} :-")
    print("\n|" + ("-" * 104) + "|")
    print(f"|{'Semester Marks':^20}|{'Hobbies':^20}|{'Cities Visited':^20}"
          f"|{'Programming Skills':^20}|{'Domain Areas':^20}|")
    for index in range(size_of_list):
        marks = marks_list[index]
        hobby = hobbies_list[index]
        city_visited = cities_visited_list[index]
        programming_skill = programming_skills_list[index]
        domain_area = domain_ares_list[index]
        print(f"|{marks:^20}|{hobby:^20}|{city_visited:^20}|"
              f"{programming_skill:^20}|{domain_area:^20}|")

    print("|" + ("-" * 104) + "|")


def write_marks_to_excel(output_sheet, column_name, list_of_marks):
    """Function to write the marks of a PS number to the excel file"""
    row_num = 2
    for mark in list_of_marks:
        cell_pos = column_name + str(row_num)
        output_sheet[cell_pos] = mark
        row_num += 1
    return True


def write_hobbies_to_excel(output_sheet, column_name, list_of_hobbies):
    """Function to write the hobbies of a PS number to the excel file"""
    row_num = 2
    for hobby in list_of_hobbies:
        cell_pos = column_name + str(row_num)
        output_sheet[cell_pos] = hobby
        row_num += 1
    return True


def write_cities_visited_to_excel(output_sheet, column_name, list_of_cities):
    """Function to write the cities of a PS number to the excel file"""
    row_num = 2
    for city in list_of_cities:
        cell_pos = column_name + str(row_num)
        output_sheet[cell_pos] = city
        row_num += 1
    return True


def write_programming_skills_to_excel(output_sheet, column_name, list_of_programming_skills):
    """Function to write the programming skills of a PS number to the excel file"""
    row_num = 2
    for programming_skill in list_of_programming_skills:
        cell_pos = column_name + str(row_num)
        output_sheet[cell_pos] = programming_skill
        row_num += 1
    return True


def write_domain_areas_to_excel(output_sheet, column_name, list_of_domains):
    """Function to write the domain areas of a PS number to the excel file"""
    row_num = 2
    for domain_area in list_of_domains:
        cell_pos = column_name + str(row_num)
        output_sheet[cell_pos] = domain_area
        row_num += 1
    return True


def write_data_to_excel(ps_number):
    """Function to write all extracted data to the output excel file
    using other functions"""
    output_workbook = openpyxl.Workbook()
    output_workbook.remove(output_workbook['Sheet'])
    output_workbook.create_sheet("Output Sheet")
    sheet = output_workbook["Output Sheet"]
    sheet["A1"] = "Semester Marks"
    sheet["B1"] = "Hobbies"
    sheet["C1"] = "Cities Visited"
    sheet["D1"] = "Programming Skills"
    sheet["E1"] = "Domain Areas"

    write_marks_to_excel(sheet, "A", marks_list)
    write_hobbies_to_excel(sheet, "B", hobbies_list)
    write_cities_visited_to_excel(sheet, "C", cities_visited_list)
    write_programming_skills_to_excel(sheet, "D", programming_skills_list)
    write_domain_areas_to_excel(sheet, "E", domain_ares_list)
    output_workbook.save(f"{ps_number}_output_data.xlsx")
    print(f"\nOutput Written to file named '{ps_number}_output_data.xlsx'")
    return True


if __name__ == '__main__':
    while True:
        try:
            input_workbook = openpyxl.open("datafile.xlsx")
            ps_numbers_list = get_ps_numbers(input_workbook)
            ps_number_choice = print_welcome_screen(ps_numbers_list)
            ps_num = ps_numbers_list[ps_number_choice - 2]
            marks_list = get_marks_ps_number(ps_number_choice, input_workbook)
            hobbies_list = get_hobbies_ps_number(ps_number_choice, input_workbook)
            cities_visited_list = get_cities_ps_number(ps_number_choice, input_workbook)
            programming_skills_list = get_programming_ps_number(ps_number_choice, input_workbook)
            domain_ares_list = get_domain_areas_ps_number(ps_number_choice, input_workbook)
            print_data_of_ps_number(ps_num, len(marks_list))
            write_data_to_excel(ps_num)
        except ValueError:
            print("ERROR!! Enter only Integer")
            print("Enter a Integer number between 1-15")
        except FileNotFoundError as error:
            print(error)
            print("Please paste Python and Excel file in same directory")
            print("Rename the file to 'datafile.xlsx' if already in directory")
        except TypeError as error:
            print("\nERROR!!", error)
            print("Some values missing in the Excel File\n")
